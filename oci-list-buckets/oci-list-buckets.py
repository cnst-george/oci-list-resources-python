import oci
import json
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import PieChart, BarChart, Reference

# Load OCI configuration
config = oci.config.from_file("~/.oci/config")

# Initialize OCI clients
identity_client = oci.identity.IdentityClient(config)
object_storage_client = oci.object_storage.ObjectStorageClient(config)

# Get Object Storage namespace
namespace = object_storage_client.get_namespace().data

# Get tenancy ID
tenancy_ocid = config["tenancy"]

# Fetch availability domains
availability_domains = identity_client.list_availability_domains(tenancy_ocid).data

# Initialize result storage
resources = {}
findings = {}

try:
    # Fetch all compartments
    compartments = oci.pagination.list_call_get_all_results(
        identity_client.list_compartments,
        tenancy_ocid,
        compartment_id_in_subtree=True,
        access_level="ANY"
    ).data
    compartments.append(oci.identity.models.Compartment(id=tenancy_ocid, name="Tenancy Root"))

    # Discover resources in each compartment
    for compartment in compartments:
        if compartment.lifecycle_state == "ACTIVE":
            print(f"Discovering resources in compartment: {compartment.name}")
            resources[compartment.name] = {}
            findings[compartment.name] = []
            
            # Discover Object Storage Buckets
            bucket_response = oci.pagination.list_call_get_all_results(
                object_storage_client.list_buckets,
                namespace_name=namespace,
                compartment_id=compartment.id
            ).data
            bucket_findings = []
            for bucket in bucket_response:
                resources[compartment.name].setdefault("Buckets", []).append({"name": bucket.name})
                # Fetch detailed bucket info to check for public access
                bucket_details = object_storage_client.get_bucket(
                    namespace_name=namespace,
                    bucket_name=bucket.name
                ).data
                # Best practice: Check for public access
                if bucket_details.public_access_type != "NoPublicAccess":
                    bucket_findings.append(f"Bucket '{bucket.name}' allows public access.")
                # Discover Objects in Buckets
                object_response = oci.pagination.list_call_get_all_results(
                    object_storage_client.list_objects,
                    namespace_name=namespace,
                    bucket_name=bucket.name
                ).data
                resources[compartment.name].setdefault("Bucket Objects", []).extend([
                    {"bucket_name": bucket.name, "object_name": obj.name} for obj in object_response.objects
                ])
            findings[compartment.name].extend(bucket_findings)

    # Get current date for the file name
    current_date = datetime.now().strftime("%Y-%m-%d")
    tenancy_name = tenancy_ocid.split(".")[1] if tenancy_ocid else "unknown"
    
    # Generate file name with dynamic titles
    file_name = f"oci_buckets_{tenancy_name}_{current_date}.json"
    
    # Export data to JSON
    with open(file_name, "w") as file:
        json.dump({"resources": resources, "findings": findings}, file, indent=4)

    print(f"JSON file saved: {file_name}")
    
    # Export data to Excel
    workbook = Workbook()
    summary_sheet = workbook.active

    # Remove default sheet
    workbook.remove(workbook["Sheet"])

    # Add data sheets for each resource type
    for resource_type in ["Buckets"
                         ,"Bucket Objects"
                        ]:
        sheet = workbook.create_sheet(title=resource_type)
        sheet.append(["Compartment", "Name", "ID"])
        for compartment, resource_data in resources.items():
            for item in resource_data.get(resource_type, []):
                sheet.append([compartment, item.get("name"), item.get("id", "N/A")])

      # Generate file name with dynamic titles
    file_name = f"oci_buckets_{tenancy_name}_{current_date}.xlsx"
    
    # Save the Excel workbook
    workbook.save(file_name)
    print(f"Excel file saved: {file_name}")

except oci.exceptions.ServiceError as e:
    print(f"Service Error: {e}")
except Exception as e:
    print(f"Unexpected Error: {e}")