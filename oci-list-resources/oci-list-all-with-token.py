import oci
import json
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import PieChart, BarChart, Reference

# Pre-requisites 
# Step.1 (required) Run:
#        oci session authenticate
# Step.2 (optional) Refresh Token:
#        oci session refresh --profile 'token'

config = oci.config.from_file(profile_name='token2')
token_file = config['security_token_file']
token = None
with open(token_file, 'r') as f:
     token = f.read()
private_key = oci.signer.load_private_key_from_file(config['key_file'])
signer = oci.auth.signers.SecurityTokenSigner(token, private_key)
region = config['region']

# Initialize OCI clients
identity_client = oci.identity.IdentityClient({'region': region}, signer=signer)
virtual_network_client = oci.core.VirtualNetworkClient({'region': region}, signer=signer)
compute_client = oci.core.ComputeClient({'region': region}, signer=signer)
block_storage_client = oci.core.BlockstorageClient({'region': region}, signer=signer)
file_storage_client = oci.file_storage.FileStorageClient({'region': region}, signer=signer)
object_storage_client = oci.object_storage.ObjectStorageClient({'region': region}, signer=signer)
database_client = oci.database.DatabaseClient({'region': region}, signer=signer)
load_balancer_client = oci.load_balancer.LoadBalancerClient({'region': region}, signer=signer)

# Get Object Storage namespace
namespace = object_storage_client.get_namespace().data

# Get tenancy ID
tenancy_ocid = config['tenancy']
tenancy_name = identity_client.get_tenancy(tenancy_id=tenancy_ocid).data.name
print(f"Using Tenancy Name: {tenancy_name}")

# Fetch availability domains
availability_domains = identity_client.list_availability_domains(tenancy_ocid).data

# Initialize result storage
resources = {}
findings = {}

try:
    # Fetch all compartments
    compartments = oci.pagination.list_call_get_all_results(
        identity_client.list_compartments,
        tenancy_ocid,
        compartment_id_in_subtree=True,
        access_level="ANY"
    ).data
    compartments.append(oci.identity.models.Compartment(id=tenancy_ocid, name="Tenancy Root"))

    # Discover resources in each compartment
    for compartment in compartments:
        if compartment.lifecycle_state == "ACTIVE":
            print(f"Discovering resources in compartment: {compartment.name}")
            resources[compartment.name] = {}
            findings[compartment.name] = []

            # Discover Compute Instances
            instance_response = oci.pagination.list_call_get_all_results(
                compute_client.list_instances,
                compartment_id=compartment.id,
            ).data
            instance_findings = []
            for instance in instance_response:
                resources[compartment.name].setdefault("Compute Instances", []).append({
                    "name": instance.display_name,
                    "id": instance.id,
                    "defined_tags" : instance.defined_tags,
                    "freeform_tags" : instance.freeform_tags
                })
                # Best practice: Check if instance metadata is restricted
                if instance.shape.startswith("VM.Standard"):
                    instance_findings.append(f"Instance '{instance.display_name}' is using '{instance.shape}' shape")
            findings[compartment.name].extend(instance_findings)

            # Discover Block Volumes
            volume_response = oci.pagination.list_call_get_all_results(
                block_storage_client.list_volumes,
                compartment_id=compartment.id
            ).data
            volume_findings = []
            for volume in volume_response:
                resources[compartment.name].setdefault("Block Volumes", []).append({
                    "name": volume.display_name,
                    "id": volume.id,
                    "defined_tags" : volume.defined_tags,
                    "freeform_tags" : volume.freeform_tags
                })
                # Check if the volume is attached to any instance 
                attachments = oci.pagination.list_call_get_all_results(
                    compute_client.list_volume_attachments,
                    compartment_id=compartment.id,
                    volume_id=volume.id
                ).data
                if not attachments:  # No attachments found
                    volume_findings.append(f"Volume '{volume.display_name}' is NOT attached to any instance.")
                # Best practice: Ensure backup policy is set
                if not volume.is_auto_tune_enabled:
                    volume_findings.append(f"Volume '{volume.display_name}' does not have auto-tune enabled.")
            findings[compartment.name].extend(volume_findings)

            # Discover File Systems 
            for ad in availability_domains:  
                fss_response = oci.pagination.list_call_get_all_results(
                    file_storage_client.list_file_systems,
                    compartment_id=compartment.id,
                    availability_domain=ad.name
                ).data
                fss_findings = []
                for fss in fss_response:
                    resources[compartment.name].setdefault("File Systems", []).append({
                        "name": fss.display_name,
                        "id": fss.id,
                        "defined_tags" : fss.defined_tags,
                        "freeform_tags" : fss.freeform_tags
                    })
                findings[compartment.name].extend(fss_findings)

            # Discover Autonomous Databases
            adb_response = oci.pagination.list_call_get_all_results(
                database_client.list_autonomous_databases,
                compartment_id=compartment.id
            ).data
            adb_findings = []
            for adb in adb_response:
                resources[compartment.name].setdefault("Autonomous Databases", []).append({
                    "name": adb.display_name,
                    "id": adb.id,
                    "defined_tags" : adb.defined_tags,
                    "freeform_tags" : adb.freeform_tags
                })
                # Best practice: Check for appropriate workload type
                if adb.db_workload != "OLTP":
                    adb_findings.append(f"ADB '{adb.display_name}' is not optimized for OLTP workloads.")
            findings[compartment.name].extend(adb_findings)

    # Get current date for the file name
    current_date = datetime.now().strftime("%Y-%m-%d")
   
    # Generate file name with dynamic titles
    file_name = f"oci_resources_{namespace}_{current_date}.json"
    
    # Export data to JSON
    with open(file_name, "w") as file:
        json.dump({"resources": resources, "findings": findings}, file, indent=4)

    print(f"Resource discovery and validation completed. Results saved to: {file_name}")
    
    # Export data to Excel
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Findings Summary"

    # Add findings summary
    summary_sheet.append(["Compartment", "Remarks"])
    for compartment, issues in findings.items():
        for issue in issues:
            summary_sheet.append([compartment, issue])

    # Style misconfigurations
    for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            cell.font = Font(bold=True)

    # Count findings by resource type for visualization
    resource_issues_summary = {}
    for compartment, issues in findings.items():
        for issue in issues:
            resource_type = issue.split(" ")[0]  # Extract resource type from the issue string
            resource_issues_summary[resource_type] = resource_issues_summary.get(resource_type, 0) + 1

    # Add a summary table for findings by resource type
    summary_start_row = summary_sheet.max_row + 2
    summary_sheet.append(["Resource Type", "Number of Issues"])
    for resource_type, count in resource_issues_summary.items():
        summary_sheet.append([resource_type, count])

    # Create a bar chart for findings summary
    bar_chart = BarChart()
    data = Reference(summary_sheet, min_col=2, min_row=summary_start_row + 1, max_row=summary_sheet.max_row)
    categories = Reference(summary_sheet, min_col=1, min_row=summary_start_row + 1, max_row=summary_sheet.max_row)
    bar_chart.add_data(data, titles_from_data=False)
    bar_chart.set_categories(categories)
    bar_chart.title = "Findings by Resource Type"
    bar_chart.x_axis.title = "Resource Type"
    bar_chart.y_axis.title = "Number of Issues"
    summary_sheet.add_chart(bar_chart, f"E{summary_start_row}")

    # Add data sheets for each resource type
    for resource_type in ["Compute Instances", 
                        "Block Volumes", 
                        "File Systems",
                        "Autonomous Databases"
                        ]:
        sheet = workbook.create_sheet(title=resource_type)
        sheet.append(["Compartment", "Name", "ID", "Defined_tags", "Freeform_tags" ])
        for compartment, resource_data in resources.items():
            for item in resource_data.get(resource_type, []):
                sheet.append([compartment, item.get("name"), item.get("id", "N/A"),str(item.get("defined_tags")),str(item.get(f"freeform_tags"))])

    # Add visualization sheet
    visualization_sheet = workbook.create_sheet(title="Visualizations")
    visualization_sheet.append(["Resource Type", "Count"])

    # Prepare summary data for visualization
    summary_data = {}
    for compartment, resource_types in resources.items():
        for resource_type, resource_list in resource_types.items():
            summary_data[resource_type] = summary_data.get(resource_type, 0) + len(resource_list)

    for resource_type, count in summary_data.items():
        visualization_sheet.append([resource_type, count])

    # Create Pie Chart
    pie_chart = PieChart()
    data = Reference(visualization_sheet, min_col=2, min_row=2, max_row=len(summary_data) + 1)
    labels = Reference(visualization_sheet, min_col=1, min_row=2, max_row=len(summary_data) + 1)
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.title = "Resource Distribution"
    visualization_sheet.add_chart(pie_chart, "D2")

    # Create Bar Chart
    bar_chart = BarChart()
    bar_chart.add_data(data, titles_from_data=False)
    bar_chart.set_categories(labels)
    bar_chart.title = "Resource Counts"
    bar_chart.x_axis.title = "Resource Type"
    bar_chart.y_axis.title = "Count"
    visualization_sheet.add_chart(bar_chart, "D20")

    # Generate file name with dynamic titles
    file_name = f"oci_resources_{namespace}_{current_date}.xlsx"
    
    # Save the Excel workbook
    workbook.save(file_name)
    print(f"Detailed findings and visualizations saved to: {file_name}")

except oci.exceptions.ServiceError as e:
    print(f"Service Error: {e}")
except Exception as e:
    print(f"Unexpected Error: {e}")