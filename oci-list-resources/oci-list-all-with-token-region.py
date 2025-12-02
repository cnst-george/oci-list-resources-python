import oci
import sys
import json
import pandas as pd
import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import PieChart, BarChart, Reference

# Pre-requisites 
# Step.1 (required) Run:
#        oci session authenticate
# Step.2 Change the profile name DEFAULT with yours
#        profile_name='DEFAULT'
# Step.3 (optional) Refresh Token:
#        oci session refresh --profile 'DEFAULT'
# Step.4 (optional) Run in multiple regions:
# python oci-list-all-with-token.py <region>
# Example:
# python oci-list-all-with-token.py eu-frankfurt-1
# python oci-list-all-with-token.py eu-zurich-1
# python oci-list-all-with-token.py eu-frankfurt-1 2025-10-01T00:00:00Z 2025-11-24T00:00:00Z
# python oci-list-all-with-token.py eu-frankfurt-1 2025-11-01T00:00:00Z 2025-11-24T00:00:00Z

configAPI = oci.config.from_file(profile_name='DEFAULT')
token_file = configAPI['security_token_file']
token = None
with open(token_file, 'r') as f:
     token = f.read()
private_key = oci.signer.load_private_key_from_file(configAPI['key_file'])
signer = oci.auth.signers.SecurityTokenSigner(token, private_key)

# Get Home Region
region = configAPI["region"] 
region_param = sys.argv[1] if len(sys.argv) > 1 else region
date_from_param = sys.argv[2] if len(sys.argv) > 2 else datetime.date.today().replace(day=1) # Get the first day of the current month
date_to_param = sys.argv[3] if len(sys.argv) > 3 else datetime.date.today() # Get the current day of the current month

# Initialize OCI clients
identity_client = oci.identity.IdentityClient({'region': region_param}, signer=signer)
compute_client = oci.core.ComputeClient({'region': region_param}, signer=signer)
block_storage_client = oci.core.BlockstorageClient({'region': region_param}, signer=signer)
file_storage_client = oci.file_storage.FileStorageClient({'region': region_param}, signer=signer)
object_storage_client = oci.object_storage.ObjectStorageClient({'region': region}, signer=signer)
database_client = oci.database.DatabaseClient({'region': region_param}, signer=signer)
usage_client = oci.usage_api.UsageapiClient({'region': region_param}, signer=signer)

# Get Object Storage namespace
namespace = object_storage_client.get_namespace().data

# Get tenancy ID
tenancy_ocid = configAPI["tenancy"]
tenancy_name = identity_client.get_tenancy(tenancy_id=tenancy_ocid).data.name
print(f"Using Tenancy Name: {tenancy_name}")

# Fetch availability domains
availability_domains = identity_client.list_availability_domains(tenancy_ocid).data
print(f"Discovering resources in Availability Domain: {availability_domains}")

# Initialize result storage
resources = {}
findings = {}
globalresources = {}

# Create the from and to dates for the usage query - using the previous calendar month
dateto = datetime.date.today().replace(day=1) # Get the first day of the current month
month, year = (dateto.month-1, dateto.year) if dateto.month != 1 else (12, dateto.year-1)
datefrom = dateto.replace(day=1, month=month, year=year) # Get the first day of the previous month

try:
    # Fetch all compartments
    cmp_list = oci.pagination.list_call_get_all_results(
        identity_client.list_compartments,
        tenancy_ocid,
        compartment_id_in_subtree=True,
        access_level="ANY"
    ).data
    cmp_list.append(oci.identity.models.Compartment(id=tenancy_ocid, name="Tenancy Root"))
    cmp_list.append(identity_client.get_compartment(tenancy_ocid).data)  # Add root compartment
    
    # Discover resources in each compartment
    for compartment in cmp_list:
        if  compartment.id.startswith("ocid1.compartment.oc1..") and compartment.id == "ocid1.compartment.oc1..aaaaaaaa64v3nqu4jauy726w3sui4r54pnbf6lphsez4e747pbbwwn3ccogq":
            print(f"Discovering resources in compartment: {compartment.name}")
            resources[compartment.id] = {}
            findings[compartment.id] = []

 # Iterate over each region and retrieve the volume details
# for region in region_response.data:
#    region_name = region.region_name
#     print(f"Switching to region: {region_name}")

            # Compute Instances
            for ad in availability_domains:
             vm_list = oci.pagination.list_call_get_all_results(
                compute_client.list_instances,
                compartment_id=compartment.id,
                availability_domain=ad.name
             ).data
             bv_attachments = oci.pagination.list_call_get_all_results(
                compute_client.list_boot_volume_attachments,
                compartment_id=compartment.id,
                availability_domain=ad.name
             ).data   
             vm_findings = []        
             for vm in vm_list:        
                for bva in bv_attachments:
                    if vm.id == bva.instance_id and region_param.upper() != "AP-TOKYO-1":
                        resources[compartment.id].setdefault("Compute Instances", []).append({
                            "compartment_name": compartment.name,
                            "name": vm.display_name,
                            "id": vm.id,
                            "state": vm.lifecycle_state,
                            "attached_to" : bva.boot_volume_id,
                            "volume_state": bva.lifecycle_state,
                            "availability_domain" : vm.availability_domain,
                            "defined_tags" : vm.defined_tags,
                            "freeform_tags" : vm.freeform_tags,
                            "time_created" : str((f"{vm.time_created}"))
                            # "shape" : vm.shape,
                            # "ocpus" : vm.shape_config.ocpus,
                            # "memory_in_gbs" : vm.shape_config.memory_in_gbs,
                            # "processor_description" : vm.shape_config.processor_description
                        })                   
            findings[compartment.id].extend(vm_findings)

            # Block Volumes
            bv_list = oci.pagination.list_call_get_all_results(
                block_storage_client.list_volumes,
                compartment_id=compartment.id
            ).data
            bv_attachments = oci.pagination.list_call_get_all_results(
                compute_client.list_volume_attachments,
                compartment_id=compartment.id
            ).data 
            bv_findings = []           
            for bv in bv_list:
                resources[compartment.id].setdefault("Block Volumes", []).append({
                    "compartment_name": compartment.name,
                    "name": bv.display_name,
                    "id": bv.id,
                    "state": bv.lifecycle_state,
                    "defined_tags" : bv.defined_tags,
                    "freeform_tags" : bv.freeform_tags,
                    "size_in_gbs" : bv.size_in_gbs,
                    "time_created" : str((f"{bv.time_created}"))
                })
                for bva in bv_attachments:
                    if bv.id == bva.volume_id and bv.id:
                        resources[compartment.id].setdefault("Block Volumes", []).append({
                            "compartment_name": compartment.name,
                            "name": bv.display_name,
                            "id": bv.id,
                            "state": bva.lifecycle_state,
                            "defined_tags" : bv.defined_tags,
                            "freeform_tags" : bv.freeform_tags,
                            "attached_to_instance" : bva.instance_id,
                            "size_in_gbs" : bv.size_in_gbs,
                            "time_created" : str((f"{bv.time_created}"))
                        })                       
            findings[compartment.id].extend(bv_findings)

            # Block Volumes Bkp
            bvBkp_list = oci.pagination.list_call_get_all_results(
                block_storage_client.list_volume_backups,
                compartment_id=compartment.id
            ).data
            bv_findings = []
            for bv in bvBkp_list:
                resources[compartment.id].setdefault("Block Volumes Bkp", []).append({
                    "compartment_name": compartment.name,
                    "name": bv.display_name,
                    "id": bv.id,
                    "state": bv.lifecycle_state,
                    "defined_tags" : bv.defined_tags,
                    "freeform_tags" : bv.freeform_tags,
                    "size_in_gbs" : bv.size_in_gbs,
                    "time_created" : str((f"{bv.time_created}"))
                })
            findings[compartment.id].extend(bv_findings)

        # Boot Volumes
            for ad in availability_domains:
             bv_list = oci.pagination.list_call_get_all_results(
                block_storage_client.list_boot_volumes,
                compartment_id=compartment.id,
                availability_domain=ad.name
             ).data
             bv_attachments = oci.pagination.list_call_get_all_results(
                compute_client.list_boot_volume_attachments,
                compartment_id=compartment.id,
                availability_domain=ad.name
             ).data   
             bv_findings = []        
             for bv in bv_list:
                resources[compartment.id].setdefault("Block Volumes", []).append({
                    "compartment_name": compartment.name,
                    "name": bv.display_name,
                    "id": bv.id,
                    "state": bv.lifecycle_state,
                    "defined_tags" : bv.defined_tags,
                    "freeform_tags" : bv.freeform_tags,
                    "size_in_gbs" : bv.size_in_gbs,
                    "time_created" : str((f"{bv.time_created}"))
                })          
                for bva in bv_attachments:
                    if bv.id == bva.boot_volume_id:
                        resources[compartment.id].setdefault("Boot Volumes", []).append({
                            "compartment_name": compartment.name,
                            "name": bv.display_name,
                            "id": bv.id,
                            "state": bva.lifecycle_state,
                            "defined_tags" : bv.defined_tags,
                            "freeform_tags" : bv.freeform_tags,
                            "attached_to_instance" : bva.instance_id,
                            "availability_domain" : ad.name,
                            "size_in_gbs" : bv.size_in_gbs,
                            "time_created" : str((f"{bv.time_created}"))
                        })                       
            findings[compartment.id].extend(bv_findings)

            # Boot Volumes Bkp
            bvBkp_list = oci.pagination.list_call_get_all_results(
                block_storage_client.list_boot_volume_backups,
                compartment_id=compartment.id
            ).data
            bv_findings = []
            for bv in bvBkp_list:
                resources[compartment.id].setdefault("Boot Volumes Bkp", []).append({
                    "compartment_name": compartment.name,
                    "name": bv.display_name,
                    "id": bv.id,
                    "state": bv.lifecycle_state,
                    "defined_tags" : bv.defined_tags,
                    "freeform_tags" : bv.freeform_tags,
                    "size_in_gbs" : bv.size_in_gbs,
                    "time_created" : str((f"{bv.time_created}"))
                })
            findings[compartment.id].extend(bv_findings)

            # File Systems 
            for ad in availability_domains:  
                fss_list = oci.pagination.list_call_get_all_results(
                    file_storage_client.list_file_systems,
                    compartment_id=compartment.id,
                    availability_domain=ad.name
                ).data
                fss_findings = []
                for fss in fss_list:
                    resources[compartment.id].setdefault("File Systems", []).append({
                        "compartment_name": compartment.name,
                        "name": fss.display_name,
                        "id": fss.id,
                        "state": fss.lifecycle_state,
                        "defined_tags" : fss.defined_tags,
                        "freeform_tags" : fss.freeform_tags,
                        "metered_bytes" : fss.metered_bytes, # (1024 * 1024 * 1024)
                        "time_created" : str((f"{fss.time_created}"))
                    })
                findings[compartment.id].extend(fss_findings)

            # Autonomous Databases
            adb_list = oci.pagination.list_call_get_all_results(
                database_client.list_autonomous_databases,
                compartment_id=compartment.id
            ).data
            adb_findings = []
            for adb in adb_list:
                resources[compartment.id].setdefault("Autonomous Databases", []).append({
                    "compartment_name": compartment.name,
                    "name": adb.display_name,
                    "id": adb.id,
                    "state": adb.lifecycle_state,
                    "defined_tags" : adb.defined_tags,
                    "freeform_tags" : adb.freeform_tags,
                    "ocups": adb.compute_count,
                    "size_in_gbs" : adb.data_storage_size_in_gbs,
                    "time_created" : str((f"{adb.time_created}"))
                })
            findings[compartment.id].extend(adb_findings)

            # Usage Costs - Compute

            # filter_details = oci.usage_api.models.Filter(
            # operator="AND",
            # dimensions=[
            #     oci.usage_api.models.Dimension(
            #         key="compartmentId",
            #         value=compartment.id
            #     ),
            #     oci.usage_api.models.Dimension(
            #         key="service",
            #         value="COMPUTE"
            #     )
            # ]
            # )

        if compartment.id.startswith("ocid1.tenancy.oc1..") and region_param.upper() == "EU-FRANKFURT-1": 
            print(f"Discovering Costs in Root Compartment: {compartment.name}")
            print(f"Date from: {date_from_param} to Date to: {date_to_param}")
            print(f"Home region: {region_param}")
            # print(f"Date from: {datefrom} to Date to: {dateto }")
            resources[compartment.id] = {}
            findings[compartment.id] = []

            # Enable debug logging
            #oci.base_client.is_http_log_enabled(True)

            # usage_list = oci.pagination.list_call_get_all_results(
            costs_list = usage_client.request_summarized_usages(
                request_summarized_usages_details=oci.usage_api.models.RequestSummarizedUsagesDetails(
                # compartment_id=compartment.id,
                tenant_id=tenancy_ocid,
                time_usage_started=date_from_param,
                time_usage_ended=date_to_param,
                # time_usage_started=(datefrom.strftime('%Y-%m-%dT%H:%M:%SZ')),
                # time_usage_ended=(dateto.strftime('%Y-%m-%dT%H:%M:%SZ')),
                granularity="DAILY",
                # filter=filter_details,
                is_aggregate_by_time=False,
                query_type="COST",
                group_by=["resourceId"],
                # group_by_tag=[
                #     oci.usage_api.models.Tag( # Return results by the CreatedBy tag, which will indicate the user who created the resource (who the usage cost will be attributed to)
                #         namespace="Oracle-Tags",
                #         key="CreatedBy")],
                # compartment_depth=1
                compartment_depth=6              
                )
            )

            cost_findings = []

            for cost in costs_list.data.items:
                start_time = cost.time_usage_started.strftime("%Y-%m-%d")
                resources[compartment.id].setdefault("Daily Costs", []).append({
                    "compartment_name": compartment.name,
                    "id": cost.resource_id,
                    "currency": cost.currency,
                    "cost": cost.computed_amount,
                    "starttime": start_time
                })
            findings[compartment.id].extend(cost_findings)
        
    # Get current date for the file name
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
   
    # Generate file name with dynamic titles
    file_name = f"oci_resources_{region_param}_{namespace}_{current_date}.json"
    
    # # Export data to JSON
    # with open(file_name, "w") as file:
    #     json.dump({"resources": resources, "findings": findings}, file, indent=4)

    # print(f"Resource discovery and validation completed. Results saved to: {file_name}")
    
    # # Export data to Excel
    workbook = Workbook()
    # summary_sheet = workbook.active
    # summary_sheet.title = "Findings Summary"

    # # Add findings summary
    # summary_sheet.append(["Compartment", "Remarks"])
    # for compartment, issues in findings.items():
    #     for issue in issues:
    #         summary_sheet.append([compartment, issue])

    # Style misconfigurations
    # for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row, min_col=2, max_col=2):
    #     for cell in row:
    #         cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    #         cell.font = Font(bold=True)

    # # Count findings by resource type for visualization
    # resource_issues_summary = {}
    # for compartment, issues in findings.items():
    #     for issue in issues:
    #         resource_type = issue.split(" ")[0]  # Extract resource type from the issue string
    #         resource_issues_summary[resource_type] = resource_issues_summary.get(resource_type, 0) + 1

    # # Add a summary table for findings by resource type
    # summary_start_row = summary_sheet.max_row + 2
    # summary_sheet.append(["Resource Type", "Number of Issues"])
    # for resource_type, count in resource_issues_summary.items():
    #     summary_sheet.append([resource_type, count])

    # Create a bar chart for findings summary
    # bar_chart = BarChart()
    # data = Reference(summary_sheet, min_col=2, min_row=summary_start_row + 1, max_row=summary_sheet.max_row)
    # categories = Reference(summary_sheet, min_col=1, min_row=summary_start_row + 1, max_row=summary_sheet.max_row)
    # bar_chart.add_data(data, titles_from_data=False)
    # bar_chart.set_categories(categories)
    # bar_chart.title = "Findings by Resource Type"
    # bar_chart.x_axis.title = "Resource Type"
    # bar_chart.y_axis.title = "Number of Issues"
    # summary_sheet.add_chart(bar_chart, f"E{summary_start_row}")

    # Add data sheets for each resource type
    for resource_type in ["Daily Costs",
                        "Compute Instances", 
                        "Block Volumes", 
                        "Block Volumes Bkp", 
                        "Boot Volumes",
                        "Boot Volumes Bkp",
                        "File Systems",
                        "Autonomous Databases"]:
        sheet = workbook.create_sheet(title=resource_type)   
        if resource_type == "Compute Instances":
            sheet.append(["Compartment", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags", "Attached_to", "BootVolume_state", "Availability_domain", "Time_created"])
        if resource_type == "Block Volumes":
            sheet.append(["Compartment", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags", "Attached_to", "Size_in_gbs", "Time_created"])
        if resource_type == "Block Volumes Bkp":
            sheet.append(["Compartment", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags" , "Size_in_gbs", "Time_created"])
        if resource_type == "Boot Volumes":
            sheet.append(["Compartment", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags", "Attached_to", "Availability_domain" , "Size_in_gbs", "Time_created"])
        if resource_type == "Boot Volumes Bkp":
            sheet.append(["Compartment", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags", "Size_in_gbs", "Time_created" ])
        if resource_type == "File Systems":
            sheet.append(["Compartment", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags" , "Metered_bytes", "Time_created"])
        if resource_type == "Autonomous Databases":
            sheet.append(["Compartment", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags" , "Ocpus", "Size_in_gbs", "Time_created"])    
        if resource_type == "Daily Costs":
            sheet.append(["Compartment", "ID", "Currency","Cost", "Starttime"])              
        
        for compartment, resource_data in resources.items():
            for item in resource_data.get(resource_type, []):           
                if resource_type == "Compute Instances":
                    sheet.append([#compartment, 
                             item.get("compartment_name"), 
                             item.get("name"), 
                             item.get("id"),
                             item.get("state"),
                             str(item.get(f"defined_tags")),
                             str(item.get(f"freeform_tags")),
                             str(item.get(f"attached_to")),
                             str(item.get(f"volume_state")),
                             str(item.get(f"availability_domain")),
                             str(item.get(f"time_created"))
                            #  str(item.get(f"shape")),
                            #  item.get(f"ocpus"),
                            #  item.get(f"memory_in_gbs"),
                            #  str(item.get(f"processor_description"))
                             ])
                if resource_type == "Block Volumes":
                    sheet.append([#compartment, 
                                item.get("compartment_name"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                str(item.get(f"attached_to_instance")),
                                item.get(f"size_in_gbs"),
                                str(item.get(f"time_created"))
                                ])
                if resource_type == "Block Volumes Bkp":
                    sheet.append([#compartment, 
                                item.get("compartment_name"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                item.get(f"size_in_gbs"),
                                str(item.get(f"time_created"))
                                ])
                if resource_type == "Boot Volumes":
                    sheet.append([#compartment,
                                item.get("compartment_name"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                str(item.get(f"attached_to_instance")),
                                str(item.get(f"availability_domain")),
                                item.get(f"size_in_gbs"),
                                str(item.get(f"time_created"))
                                ])
                if resource_type == "Boot Volumes Bkp":
                    sheet.append([#compartment,
                                item.get("compartment_name"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                item.get(f"size_in_gbs"),
                                str(item.get(f"time_created"))
                                ])
                if resource_type == "File Systems":
                    sheet.append([#compartment,
                                item.get("compartment_name"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                item.get(f"metered_bytes"),
                                str(item.get(f"time_created"))
                                ])
                if resource_type == "Autonomous Databases":
                    sheet.append([#compartment, 
                                item.get("compartment_name"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                item.get(f"compute_count"),
                                item.get(f"size_in_gbs"),
                                str(item.get(f"time_created"))
                                ])
                if resource_type == "Daily Costs":
                    sheet.append([#compartment, 
                                item.get("compartment_name"), 
                                item.get("id"),
                                item.get("currency"),
                                item.get("cost"),
                                item.get("starttime")
                                ])                        

    # Add visualization sheet
    visualization_sheet = workbook.create_sheet(title="Visualizations")
    visualization_sheet.append(["Resource Type", "Count"])

    # Prepare summary data for visualization
    summary_data = {}
    for compartment, resource_types in resources.items():
        for resource_type, resource_list in resource_types.items():
          if resource_type != "Daily Costs":  
            summary_data[resource_type] = summary_data.get(resource_type, 0) + len(resource_list)

    for resource_type, count in summary_data.items():
      if resource_type != "Daily Costs": 
        visualization_sheet.append([resource_type, count])

    # Create Pie Chart
    pie_chart = PieChart()
    data = Reference(visualization_sheet, min_col=2, min_row=2, max_row=len(summary_data) + 1)
    labels = Reference(visualization_sheet, min_col=1, min_row=2, max_row=len(summary_data) + 1)
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.title = "Resource Distribution"
    visualization_sheet.add_chart(pie_chart, "D2")

    # Create Bar Chart
    bar_chart = BarChart()
    bar_chart.add_data(data, titles_from_data=False)
    bar_chart.set_categories(labels)
    bar_chart.title = "Resource Counts"
    bar_chart.x_axis.title = "Resource Type"
    bar_chart.y_axis.title = "Count"
    visualization_sheet.add_chart(bar_chart, "D20")

    # Generate file name with dynamic titles
    file_name = f"oci_resources_{region_param}_{namespace}_{current_date}.xlsx"
    
    # Save the Excel workbook
    workbook.save(file_name)
    print(f"Detailed findings and visualizations saved to: {file_name}")


except oci.exceptions.ServiceError as e:
    print(f"Service Error: {e}")
except Exception as e:
    print(f"Unexpected Error: {e}")