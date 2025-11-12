import oci
import json
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import PieChart, BarChart, Reference

# Load OCI configuration
configAPI = oci.config.from_file("~/.oci/config")

# Initialize OCI clients
identity_client = oci.identity.IdentityClient(configAPI)
virtual_network_client = oci.core.VirtualNetworkClient(configAPI)
compute_client = oci.core.ComputeClient(configAPI)
block_storage_client = oci.core.BlockstorageClient(configAPI)
file_storage_client = oci.file_storage.FileStorageClient(configAPI)
object_storage_client = oci.object_storage.ObjectStorageClient(configAPI)
database_client = oci.database.DatabaseClient(configAPI)
load_balancer_client = oci.load_balancer.LoadBalancerClient(configAPI)

# Get Object Storage namespace
namespace = object_storage_client.get_namespace().data

# Get tenancy ID
tenancy_ocid = configAPI["tenancy"]
# tenancy_name = tenancy_ocid.split(".")[1] if tenancy_ocid else "unknown"
tenancy_name = identity_client.get_tenancy(tenancy_id=tenancy_ocid).data.name
print(f"Using Tenancy Name: {tenancy_name}")

# Fetch availability domains
availability_domains = identity_client.list_availability_domains(tenancy_ocid).data
print(f"Discovering resources in Availability Domain: {availability_domains}")

# Initialize result storage
resources = {}
findings = {}

try:
    # Fetch all compartments
    compartments = oci.pagination.list_call_get_all_results(
        identity_client.list_compartments,
        tenancy_ocid,
        compartment_id_in_subtree=True,
        access_level="ANY"
    ).data
    compartments.append(oci.identity.models.Compartment(id=tenancy_ocid, name="Tenancy Root"))

    # Discover resources in each compartment
    for compartment in compartments:
        if compartment.lifecycle_state == "ACTIVE":
            print(f"Discovering resources in compartment: {compartment.name}")
            resources[compartment.name] = {}
            findings[compartment.name] = []

            # Compute Instances
            vm_list = oci.pagination.list_call_get_all_results(
                compute_client.list_instances,
                compartment_id=compartment.id,
            ).data
            instance_findings = []
            for vm in vm_list:
                resources[compartment.name].setdefault("Compute Instances", []).append({
                    "name": vm.display_name,
                    "id": vm.id,
                    "defined_tags" : vm.defined_tags,
                    "freeform_tags" : vm.freeform_tags
                })
                # Best practice: Check if instance metadata is restricted
                if vm.shape.startswith("VM.Standard"):
                    instance_findings.append(f"Instance '{vm.display_name}' is using '{vm.shape}' shape")
            findings[compartment.name].extend(instance_findings)

            # Block Volumes
            bv_list = oci.pagination.list_call_get_all_results(
                block_storage_client.list_volumes,
                compartment_id=compartment.id,
                lifecycle_state='AVAILABLE'
            ).data
            bv_attachments = oci.pagination.list_call_get_all_results(
                compute_client.list_volume_attachments,
                compartment_id=compartment.id
            ).data 
            bv_findings = []           
            for bv in bv_list:
                resources[compartment.name].setdefault("Block Volumes", []).append({
                    "name": bv.display_name,
                    "id": bv.id,
                    "defined_tags" : bv.defined_tags,
                    "freeform_tags" : bv.freeform_tags
                })
                for bva in bv_attachments:
                    if bv.id == bva.volume_id:
                        resources[compartment.name].setdefault("Block Volumes", []).append({
                            "name": bv.display_name,
                            "id": bv.id,
                            "defined_tags" : bv.defined_tags,
                            "freeform_tags" : bv.freeform_tags,
                            "attached_to_instance" : bva.instance_id
                        })
                        bv_findings.append(f"Block Volume '{bv.display_name}={bv.id}' is ' {bva.lifecycle_state}' to instance' {bva.instance_id}")                         
            findings[compartment.name].extend(bv_findings)

            # Block Volumes Bkp
            bvBkp_list = oci.pagination.list_call_get_all_results(
                block_storage_client.list_volume_backups,
                compartment_id=compartment.id
            ).data
            bv_findings = []
            for bv in bvBkp_list:
                resources[compartment.name].setdefault("Block Volumes Bkp", []).append({
                    "name": bv.display_name,
                    "id": bv.id,
                    "defined_tags" : bv.defined_tags,
                    "freeform_tags" : bv.freeform_tags
                })
            findings[compartment.name].extend(bv_findings)

        # Boot Volumes
            for ad in availability_domains:
             bv_list = oci.pagination.list_call_get_all_results(
                block_storage_client.list_boot_volumes,
                compartment_id=compartment.id,
                availability_domain=ad.name
             ).data
             bv_attachments = oci.pagination.list_call_get_all_results(
                compute_client.list_boot_volume_attachments,
                compartment_id=compartment.id,
                availability_domain=ad.name
             ).data   
             bv_findings = []        
             for bv in bv_list:
                resources[compartment.name].setdefault("Block Volumes", []).append({
                    "name": bv.display_name,
                    "id": bv.id,
                    "defined_tags" : bv.defined_tags,
                    "freeform_tags" : bv.freeform_tags
                })           
                for bva in bv_attachments:
                    if bv.id == bva.boot_volume_id:
                        resources[compartment.name].setdefault("Boot Volumes", []).append({
                            "name": bv.display_name,
                            "id": bv.id,
                            "defined_tags" : bv.defined_tags,
                            "freeform_tags" : bv.freeform_tags,
                            "attached_to_instance" : bva.instance_id
                        })
                        bv_findings.append(f"Boot Volume '{bv.display_name}={bv.id}' is ' {bva.lifecycle_state}' to instance' {bva.instance_id}")                         
            findings[compartment.name].extend(bv_findings)

            # Boot Volumes Bkp
            bvBkp_list = oci.pagination.list_call_get_all_results(
                block_storage_client.list_boot_volume_backups,
                compartment_id=compartment.id
            ).data
            bv_findings = []
            for bv in bvBkp_list:
                resources[compartment.name].setdefault("Boot Volumes Bkp", []).append({
                    "name": bv.display_name,
                    "id": bv.id,
                    "defined_tags" : bv.defined_tags,
                    "freeform_tags" : bv.freeform_tags
                })
            findings[compartment.name].extend(bv_findings)

            # File Systems 
            for ad in availability_domains:  
                fss_list = oci.pagination.list_call_get_all_results(
                    file_storage_client.list_file_systems,
                    compartment_id=compartment.id,
                    availability_domain=ad.name
                ).data
                fss_findings = []
                for fss in fss_list:
                    resources[compartment.name].setdefault("File Systems", []).append({
                        "name": fss.display_name,
                        "id": fss.id,
                        "defined_tags" : fss.defined_tags,
                        "freeform_tags" : fss.freeform_tags
                    })
                findings[compartment.name].extend(fss_findings)

            # Autonomous Databases
            adb_list = oci.pagination.list_call_get_all_results(
                database_client.list_autonomous_databases,
                compartment_id=compartment.id
            ).data
            adb_findings = []
            for adb in adb_list:
                resources[compartment.name].setdefault("Autonomous Databases", []).append({
                    "name": adb.display_name,
                    "id": adb.id,
                    "defined_tags" : adb.defined_tags,
                    "freeform_tags" : adb.freeform_tags
                })
                if adb.db_workload != "OLTP":
                    adb_findings.append(f"ADB '{adb.display_name}' is not optimized for OLTP workloads.")
            findings[compartment.name].extend(adb_findings)

    # Get current date for the file name
    current_date = datetime.now().strftime("%Y-%m-%d")
   
    # Generate file name with dynamic titles
    file_name = f"oci_resources_{namespace}_{current_date}.json"
    
    # Export data to JSON
    with open(file_name, "w") as file:
        json.dump({"resources": resources, "findings": findings}, file, indent=4)

    print(f"Resource discovery and validation completed. Results saved to: {file_name}")
    
    # Export data to Excel
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Findings Summary"

    # Add findings summary
    summary_sheet.append(["Compartment", "Remarks"])
    for compartment, issues in findings.items():
        for issue in issues:
            summary_sheet.append([compartment, issue])

    # Style misconfigurations
    for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            cell.font = Font(bold=True)

    # Count findings by resource type for visualization
    resource_issues_summary = {}
    for compartment, issues in findings.items():
        for issue in issues:
            resource_type = issue.split(" ")[0]  # Extract resource type from the issue string
            resource_issues_summary[resource_type] = resource_issues_summary.get(resource_type, 0) + 1

    # Add a summary table for findings by resource type
    summary_start_row = summary_sheet.max_row + 2
    summary_sheet.append(["Resource Type", "Number of Issues"])
    for resource_type, count in resource_issues_summary.items():
        summary_sheet.append([resource_type, count])

    # Create a bar chart for findings summary
    bar_chart = BarChart()
    data = Reference(summary_sheet, min_col=2, min_row=summary_start_row + 1, max_row=summary_sheet.max_row)
    categories = Reference(summary_sheet, min_col=1, min_row=summary_start_row + 1, max_row=summary_sheet.max_row)
    bar_chart.add_data(data, titles_from_data=False)
    bar_chart.set_categories(categories)
    bar_chart.title = "Findings by Resource Type"
    bar_chart.x_axis.title = "Resource Type"
    bar_chart.y_axis.title = "Number of Issues"
    summary_sheet.add_chart(bar_chart, f"E{summary_start_row}")

    # Add data sheets for each resource type
    for resource_type in ["Compute Instances", 
                        "Block Volumes", 
                        "Block Volumes Bkp", 
                        "Boot Volumes",
                        "Boot Volumes Bkp",
                        "File Systems",
                        "Autonomous Databases"
                        ]:
        sheet = workbook.create_sheet(title=resource_type)
        sheet.append(["Compartment", "Name", "ID", "Defined_tags", "Freeform_tags", "Attached_to" ])
        for compartment, resource_data in resources.items():
            for item in resource_data.get(resource_type, []):
                sheet.append([compartment, item.get("name"), item.get("id", "N/A"),str(item.get("defined_tags")),str(item.get(f"freeform_tags")),str(item.get(f"attached_to_instance"))])

    # Add visualization sheet
    visualization_sheet = workbook.create_sheet(title="Visualizations")
    visualization_sheet.append(["Resource Type", "Count"])

    # Prepare summary data for visualization
    summary_data = {}
    for compartment, resource_types in resources.items():
        for resource_type, resource_list in resource_types.items():
            summary_data[resource_type] = summary_data.get(resource_type, 0) + len(resource_list)

    for resource_type, count in summary_data.items():
        visualization_sheet.append([resource_type, count])

    # Create Pie Chart
    pie_chart = PieChart()
    data = Reference(visualization_sheet, min_col=2, min_row=2, max_row=len(summary_data) + 1)
    labels = Reference(visualization_sheet, min_col=1, min_row=2, max_row=len(summary_data) + 1)
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.title = "Resource Distribution"
    visualization_sheet.add_chart(pie_chart, "D2")

    # Create Bar Chart
    bar_chart = BarChart()
    bar_chart.add_data(data, titles_from_data=False)
    bar_chart.set_categories(labels)
    bar_chart.title = "Resource Counts"
    bar_chart.x_axis.title = "Resource Type"
    bar_chart.y_axis.title = "Count"
    visualization_sheet.add_chart(bar_chart, "D20")

    # Generate file name with dynamic titles
    file_name = f"oci_resources_{namespace}_{current_date}.xlsx"
    
    # Save the Excel workbook
    workbook.save(file_name)
    print(f"Detailed findings and visualizations saved to: {file_name}")

except oci.exceptions.ServiceError as e:
    print(f"Service Error: {e}")
except Exception as e:
    print(f"Unexpected Error: {e}")