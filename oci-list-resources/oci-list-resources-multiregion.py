import oci
import sys
import json
import pandas as pd
import datetime
import logging
# from datetime import timedelta
# from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import PieChart, BarChart, Reference

# Load OCI configuration
configAPI = oci.config.from_file("~/.oci/config")

# Get Home Region
homeRegion = configAPI["region"] 
# region_param = sys.argv[1] if len(sys.argv) > 1 else homeRegion
date_from_param = sys.argv[1] if len(sys.argv) > 2 else datetime.date.today().replace(day=1) # Get the first day of the current month
date_to_param = sys.argv[2] if len(sys.argv) > 3 else datetime.date.today() # Get the current day of the current month

# Initialize OCI clients
identity_client = oci.identity.IdentityClient(configAPI)
# Get the list of subscribed regions
region_subscriptions = identity_client.list_region_subscriptions(tenancy_id=configAPI["tenancy"]).data
print(f"Subscribed regions: {[r.region_name for r in region_subscriptions]}")

# Get Object Storage namespace (global, doesn't need region-specific client)
object_storage_client = oci.object_storage.ObjectStorageClient(configAPI)
namespace = object_storage_client.get_namespace().data

# Initialize Usage API client for home region only (costs are only available from home region)
usage_client = oci.usage_api.UsageapiClient(configAPI)

# Get tenancy ID
tenancy_ocid = configAPI["tenancy"]
tenancy_name = identity_client.get_tenancy(tenancy_id=tenancy_ocid).data.name
print(f"Using Tenancy Name: {tenancy_name}")

# Initialize result storage
resources = {}
findings = {}
globalresources = {}

# Create the from and to dates for the usage query - using the previous calendar month
dateto = datetime.date.today().replace(day=1) # Get the first day of the current month
month, year = (dateto.month-1, dateto.year) if dateto.month != 1 else (12, dateto.year-1)
datefrom = dateto.replace(day=1, month=month, year=year) # Get the first day of the previous month
scipt_start_time = datetime.datetime.now()

try:
    # Fetch all compartments
    cmp_list = oci.pagination.list_call_get_all_results(
        identity_client.list_compartments,
        tenancy_ocid,
        compartment_id_in_subtree=True,
        access_level="ANY"
    ).data
    cmp_list.append(oci.identity.models.Compartment(id=tenancy_ocid, name=identity_client.get_compartment(tenancy_ocid).data.name)) # Add root compartment
    # cmp_list.append(identity_client.get_compartment(tenancy_ocid).data)  # Add root compartment
    
    # Iterate over each subscribed region
    for region_subscription in region_subscriptions:
    #  if region_subscription.region_name.upper() != "AP-TOKYO-1":
        current_region = region_subscription.region_name
        print(f"\n{'='*60}")
        print(f"Switching to region: {current_region}")
        print(f"{'='*60}")
        
        # Create region-specific OCI clients by setting the region on each client
        # Using the base config's signer but overriding the region
        region_identity_client = oci.identity.IdentityClient(configAPI)
        region_identity_client.base_client.set_region(current_region)
        
        virtual_network_client = oci.core.VirtualNetworkClient(configAPI)
        virtual_network_client.base_client.set_region(current_region)
        
        compute_client = oci.core.ComputeClient(configAPI)
        compute_client.base_client.set_region(current_region)
        
        block_storage_client = oci.core.BlockstorageClient(configAPI)
        block_storage_client.base_client.set_region(current_region)
        
        file_storage_client = oci.file_storage.FileStorageClient(configAPI)
        file_storage_client.base_client.set_region(current_region)
        
        database_client = oci.database.DatabaseClient(configAPI)
        database_client.base_client.set_region(current_region)
        
        load_balancer_client = oci.load_balancer.LoadBalancerClient(configAPI)
        load_balancer_client.base_client.set_region(current_region)
        # Note: usage_client is NOT region-specific - it uses home region client defined above
        
        # Fetch availability domains for this specific region using region-specific identity client
        region_ads = region_identity_client.list_availability_domains(tenancy_ocid).data
        print(f"Availability Domains in {current_region}: {[ad.name for ad in region_ads]}")
        
        # Discover resources in each compartment for this region
        for compartment in cmp_list:
            if  compartment.id.startswith("ocid1.compartment.oc1.."):
                print(f"Discovering resources in compartment: {compartment.name} (Region: {current_region})")
                # Use a composite key for resources to track region
                resource_key = f"{compartment.id}_{current_region}"
                resources[resource_key] = {}
                findings[resource_key] = []

                # Compute Instances
                for ad in region_ads:
                    vm_list = oci.pagination.list_call_get_all_results(
                        compute_client.list_instances,
                        compartment_id=compartment.id,
                        availability_domain=ad.name
                    ).data
                    bv_attachments = oci.pagination.list_call_get_all_results(
                        compute_client.list_boot_volume_attachments,
                        compartment_id=compartment.id,
                        availability_domain=ad.name
                    ).data   
                    vm_findings = []        
                    for vm in vm_list:        
                        for bva in bv_attachments:
                            if vm.id == bva.instance_id:
                                resources[resource_key].setdefault("Compute Instances", []).append({
                                    "compartment_name": compartment.name,
                                    "region": current_region,
                                    "name": vm.display_name,
                                    "id": vm.id,
                                    "state": vm.lifecycle_state,
                                    "attached_to" : bva.boot_volume_id,
                                    "volume_state": bva.lifecycle_state,
                                    "availability_domain" : vm.availability_domain,
                                    "defined_tags" : vm.defined_tags,
                                    "freeform_tags" : vm.freeform_tags,
                                    "time_created" : str((f"{vm.time_created}"))
                                })                   
                findings[resource_key].extend(vm_findings)

                # Block Volumes
                bv_list = oci.pagination.list_call_get_all_results(
                    block_storage_client.list_volumes,
                    compartment_id=compartment.id
                ).data
                bv_attachments = oci.pagination.list_call_get_all_results(
                    compute_client.list_volume_attachments,
                    compartment_id=compartment.id
                ).data 
                bv_findings = []           
                for bv in bv_list:
                    resources[resource_key].setdefault("Block Volumes", []).append({
                        "compartment_name": compartment.name,
                        "region": current_region,
                        "name": bv.display_name,
                        "id": bv.id,
                        "state": bv.lifecycle_state,
                        "defined_tags" : bv.defined_tags,
                        "freeform_tags" : bv.freeform_tags,
                        "size_in_gbs" : bv.size_in_gbs,
                        "time_created" : str((f"{bv.time_created}"))
                    })
                    for bva in bv_attachments:
                        if bv.id == bva.volume_id:
                            resources[resource_key].setdefault("Block Volumes", []).append({
                                "compartment_name": compartment.name,
                                "region": current_region,
                                "name": bv.display_name,
                                "id": bv.id,
                                "state": bva.lifecycle_state,
                                "defined_tags" : bv.defined_tags,
                                "freeform_tags" : bv.freeform_tags,
                                "attached_to_instance" : bva.instance_id,
                                "size_in_gbs" : bv.size_in_gbs,
                                "time_created" : str((f"{bv.time_created}"))
                            })                       
                findings[resource_key].extend(bv_findings)

                # Block Volumes Bkp
                bvBkp_list = oci.pagination.list_call_get_all_results(
                    block_storage_client.list_volume_backups,
                    compartment_id=compartment.id
                ).data
                bv_findings = []
                for bvBkp in bvBkp_list:
                    resources[resource_key].setdefault("Block Volumes Bkp", []).append({
                        "compartment_name": compartment.name,
                        "region": current_region,
                        "name": bvBkp.display_name,
                        "id": bvBkp.id,
                        "state": bvBkp.lifecycle_state,
                        "defined_tags" : bvBkp.defined_tags,
                        "freeform_tags" : bvBkp.freeform_tags,
                        "attached_to" : bvBkp.volume_id,
                        "size_in_gbs" : bvBkp.size_in_gbs,
                        "time_created" : str((f"{bvBkp.time_created}"))
                    })
                findings[resource_key].extend(bv_findings)

                # Boot Volumes
                for ad in region_ads:
                    bv_list = oci.pagination.list_call_get_all_results(
                        block_storage_client.list_boot_volumes,
                        compartment_id=compartment.id,
                        availability_domain=ad.name
                    ).data
                    bv_attachments = oci.pagination.list_call_get_all_results(
                        compute_client.list_boot_volume_attachments,
                        compartment_id=compartment.id,
                        availability_domain=ad.name
                    ).data   
                    bv_findings = []        
                    for bv in bv_list:
                        resources[resource_key].setdefault("Boot Volumes", []).append({
                            "compartment_name": compartment.name,
                            "region": current_region,
                            "name": bv.display_name,
                            "id": bv.id,
                            "state": bv.lifecycle_state,
                            "defined_tags" : bv.defined_tags,
                            "freeform_tags" : bv.freeform_tags,
                            "size_in_gbs" : bv.size_in_gbs,
                            "time_created" : str((f"{bv.time_created}"))
                        })          
                        for bva in bv_attachments:
                            if bv.id == bva.boot_volume_id:
                                resources[resource_key].setdefault("Boot Volumes", []).append({
                                    "compartment_name": compartment.name,
                                    "region": current_region,
                                    "name": bv.display_name,
                                    "id": bv.id,
                                    "state": bva.lifecycle_state,
                                    "defined_tags" : bv.defined_tags,
                                    "freeform_tags" : bv.freeform_tags,
                                    "attached_to_instance" : bva.instance_id,
                                    "availability_domain" : ad.name,
                                    "size_in_gbs" : bv.size_in_gbs,
                                    "time_created" : str((f"{bv.time_created}"))
                                })                       
                findings[resource_key].extend(bv_findings)

                # Boot Volumes Bkp
                bvBkp_list = oci.pagination.list_call_get_all_results(
                    block_storage_client.list_boot_volume_backups,
                    compartment_id=compartment.id
                ).data
                bv_findings = []
                for bv in bvBkp_list:
                    resources[resource_key].setdefault("Boot Volumes Bkp", []).append({
                        "compartment_name": compartment.name,
                        "region": current_region,
                        "name": bv.display_name,
                        "id": bv.id,
                        "state": bv.lifecycle_state,
                        "defined_tags" : bv.defined_tags,
                        "freeform_tags" : bv.freeform_tags,
                        "size_in_gbs" : bv.size_in_gbs,
                        "time_created" : str((f"{bv.time_created}"))
                    })
                findings[resource_key].extend(bv_findings)

                # File Systems 
                for ad in region_ads:  
                    fss_list = oci.pagination.list_call_get_all_results(
                        file_storage_client.list_file_systems,
                        compartment_id=compartment.id,
                        availability_domain=ad.name
                    ).data
                    fss_findings = []
                    for fss in fss_list:
                        resources[resource_key].setdefault("File Systems", []).append({
                            "compartment_name": compartment.name,
                            "region": current_region,
                            "name": fss.display_name,
                            "id": fss.id,
                            "state": fss.lifecycle_state,
                            "defined_tags" : fss.defined_tags,
                            "freeform_tags" : fss.freeform_tags,
                            "metered_bytes" : fss.metered_bytes,
                            "time_created" : str((f"{fss.time_created}"))
                        })
                    findings[resource_key].extend(fss_findings)

                # Autonomous Databases
                adb_list = oci.pagination.list_call_get_all_results(
                    database_client.list_autonomous_databases,
                    compartment_id=compartment.id
                ).data
                adb_findings = []
                for adb in adb_list:
                    resources[resource_key].setdefault("Autonomous Databases", []).append({
                        "compartment_name": compartment.name,
                        "region": current_region,
                        "name": adb.display_name,
                        "id": adb.id,
                        "state": adb.lifecycle_state,
                        "defined_tags" : adb.defined_tags,
                        "freeform_tags" : adb.freeform_tags,
                        "ocups": adb.compute_count,
                        "size_in_gbs" : adb.data_storage_size_in_gbs,
                        "time_created" : str((f"{adb.time_created}"))
                    })
                findings[resource_key].extend(adb_findings)

            # Usage Costs - Compute

            # filter_details = oci.usage_api.models.Filter(
            # operator="AND",
            # dimensions=[
            #     oci.usage_api.models.Dimension(
            #         key="compartmentId",
            #         value=compartment.id
            #     ),
            #     oci.usage_api.models.Dimension(
            #         key="service",
            #         value="COMPUTE"
            #     )
            # ]
            # )

        if compartment.id.startswith("ocid1.tenancy.oc1..") and current_region.upper() == homeRegion.upper(): 
            print(f"Discovering Costs in Root Compartment: {compartment.name}")
            print(f"Date from: {date_from_param} to Date to: {date_to_param}")
            print(f"Home region: {homeRegion}")
            # print(f"Date from: {datefrom} to Date to: {dateto }")
            resources[compartment.id] = {}
            findings[compartment.id] = []

            # Enable debug logging
            #oci.base_client.is_http_log_enabled(True)

            # usage_list = oci.pagination.list_call_get_all_results(
            costs_list = usage_client.request_summarized_usages(
                request_summarized_usages_details=oci.usage_api.models.RequestSummarizedUsagesDetails(
                # compartment_id=compartment.id,
                tenant_id=tenancy_ocid,
                time_usage_started=date_from_param,
                time_usage_ended=date_to_param,
                # time_usage_started=(datefrom.strftime('%Y-%m-%dT%H:%M:%SZ')),
                # time_usage_ended=(dateto.strftime('%Y-%m-%dT%H:%M:%SZ')),
                granularity="DAILY",
                # filter=filter_details,
                is_aggregate_by_time=False,
                query_type="COST",
                group_by=["resourceId"],
                # group_by_tag=[
                #     oci.usage_api.models.Tag( # Return results by the CreatedBy tag, which will indicate the user who created the resource (who the usage cost will be attributed to)
                #         namespace="Oracle-Tags",
                #         key="CreatedBy")],
                # compartment_depth=1
                compartment_depth=6              
                )
            )

            cost_findings = []

            for cost in costs_list.data.items:
                start_time = cost.time_usage_started.strftime("%Y-%m-%d")
                resources[compartment.id].setdefault("Daily Costs", []).append({
                    "compartment_name": compartment.name,
                    "id": cost.resource_id,
                    "currency": cost.currency,
                    "cost": cost.computed_amount,
                    "starttime": start_time
                })
            findings[compartment.id].extend(cost_findings)
    
    # Build a set of all existing resource IDs from all resource sheets
    existing_resource_ids = set()
    for resource_key, resource_data in resources.items():
        for resource_type, resource_list in resource_data.items():
            if resource_type != "Daily Costs":
                for item in resource_list:
                    if item.get("id"):
                        existing_resource_ids.add(item.get("id"))
    
    print(f"\nTotal existing resources found: {len(existing_resource_ids)}")
    
    # Collect all unique cost-incurring resource IDs and aggregate their total cost
    cost_resource_ids = {}
    for resource_key, resource_data in resources.items():
        for cost_item in resource_data.get("Daily Costs", []):
            resource_id = cost_item.get("id")
            if resource_id:
                if resource_id not in cost_resource_ids:
                    cost_resource_ids[resource_id] = {
                        "total_cost": 0,
                        "currency": cost_item.get("currency")
                    }
                cost_resource_ids[resource_id]["total_cost"] += cost_item.get("cost", 0) or 0
    
    print(f"Total cost-incurring resources: {len(cost_resource_ids)}")
    
    # Find resources that incurred costs but no longer exist
    deleted_resource_ids = set(cost_resource_ids.keys()) - existing_resource_ids
    # Filter out None and empty strings
    deleted_resource_ids = {rid for rid in deleted_resource_ids if rid}
    
    print(f"Deleted/Not found resources with costs: {len(deleted_resource_ids)}")
    
    # Map OCID prefixes to resource types
    ocid_to_resource_type = {
        "ocid1.instance.": "Compute Instances",
        "ocid1.volume.": "Block Volumes",
        "ocid1.volumebackup.": "Block Volumes Bkp",
        "ocid1.bootvolume.": "Boot Volumes",
        "ocid1.bootvolumebackup.": "Boot Volumes Bkp",
        "ocid1.filesystem.": "File Systems",
        "ocid1.autonomousdatabase.": "Autonomous Databases",
    }
    
    # Add deleted resources to appropriate resource sheets
    deleted_resources_key = "DELETED_RESOURCES"
    resources[deleted_resources_key] = {}
    
    for resource_id in deleted_resource_ids:
        # Determine resource type from OCID
        resource_type = "Unknown"
        for ocid_prefix, rtype in ocid_to_resource_type.items():
            if resource_id and resource_id.startswith(ocid_prefix):
                resource_type = rtype
                break
        
        # Extract region from OCID if possible (OCIDs contain region info)
        region_from_ocid = "unknown"
        if resource_id:
            # OCID format: ocid1.<resource_type>.<realm>.<region>.<unique_id>
            parts = resource_id.split(".")
            if len(parts) >= 4:
                region_from_ocid = parts[3] if parts[3] else "unknown"
        
        total_cost = cost_resource_ids.get(resource_id, {}).get("total_cost", 0)
        currency = cost_resource_ids.get(resource_id, {}).get("currency", "")
        
        # Create entry for deleted resource
        deleted_entry = {
            "compartment_name": "DELETED/NOT_FOUND",
            "region": region_from_ocid,
            "name": f"[DELETED] {resource_id}",
            "id": resource_id,
            "state": "DELETED",
            "defined_tags": {},
            "freeform_tags": {},
            "time_created": "N/A",
            "total_cost_in_period": f"{total_cost:.4f}"
        }
        
        # Add type-specific fields
        if resource_type in ["Block Volumes", "Block Volumes Bkp", "Boot Volumes", "Boot Volumes Bkp"]:
            deleted_entry["size_in_gbs"] = "N/A"
        if resource_type in ["Block Volumes Bkp"]:
            deleted_entry["attached_to"] = "N/A"
        if resource_type in ["Boot Volumes"]:
            deleted_entry["attached_to_instance"] = "N/A"
            deleted_entry["availability_domain"] = "N/A"
        if resource_type == "Compute Instances":
            deleted_entry["attached_to"] = "N/A"
            deleted_entry["volume_state"] = "N/A"
            deleted_entry["availability_domain"] = "N/A"
        if resource_type == "File Systems":
            deleted_entry["metered_bytes"] = "N/A"
        if resource_type == "Autonomous Databases":
            deleted_entry["ocups"] = "N/A"
            deleted_entry["size_in_gbs"] = "N/A"
        
        resources[deleted_resources_key].setdefault(resource_type, []).append(deleted_entry)
    
    # Count deleted resources by type
    for rtype, rlist in resources.get(deleted_resources_key, {}).items():
        print(f"  - {rtype}: {len(rlist)} deleted resources with costs")
        
    # Get current date for the file name
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
   
    # Generate file name with dynamic titles
    # file_name = f"oci_resources_{region_param}_{namespace}_{current_date}.json"
    
    # # Export data to JSON
    # with open(file_name, "w") as file:
    #     json.dump({"resources": resources, "findings": findings}, file, indent=4)

    # print(f"Resource discovery and validation completed. Results saved to: {file_name}")
    
    # # Export data to Excel
    workbook = Workbook()
    summary_sheet = workbook.active
    # summary_sheet.title = "Findings Summary"
    script_end_time = datetime.datetime.now()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["started_at", "completed_at"])
    summary_sheet.append([str((f"{scipt_start_time}")), str((f"{script_end_time}"))])

    # # Add findings summary
    # summary_sheet.append(["Compartment", "Remarks"])
    # for compartment, issues in findings.items():
    #     for issue in issues:
    #         summary_sheet.append([compartment, issue])

    # Style misconfigurations
    # for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row, min_col=2, max_col=2):
    #     for cell in row:
    #         cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    #         cell.font = Font(bold=True)

    # # Count findings by resource type for visualization
    # resource_issues_summary = {}
    # for compartment, issues in findings.items():
    #     for issue in issues:
    #         resource_type = issue.split(" ")[0]  # Extract resource type from the issue string
    #         resource_issues_summary[resource_type] = resource_issues_summary.get(resource_type, 0) + 1

    # # Add a summary table for findings by resource type
    # summary_start_row = summary_sheet.max_row + 2
    # summary_sheet.append(["Resource Type", "Number of Issues"])
    # for resource_type, count in resource_issues_summary.items():
    #     summary_sheet.append([resource_type, count])

    # Create a bar chart for findings summary
    # bar_chart = BarChart()
    # data = Reference(summary_sheet, min_col=2, min_row=summary_start_row + 1, max_row=summary_sheet.max_row)
    # categories = Reference(summary_sheet, min_col=1, min_row=summary_start_row + 1, max_row=summary_sheet.max_row)
    # bar_chart.add_data(data, titles_from_data=False)
    # bar_chart.set_categories(categories)
    # bar_chart.title = "Findings by Resource Type"
    # bar_chart.x_axis.title = "Resource Type"
    # bar_chart.y_axis.title = "Number of Issues"
    # summary_sheet.add_chart(bar_chart, f"E{summary_start_row}")

    # Add data sheets for each resource type
    for resource_type in ["Daily Costs",
                        "Compute Instances", 
                        "Block Volumes", 
                        "Block Volumes Bkp", 
                        "Boot Volumes",
                        "Boot Volumes Bkp",
                        "File Systems",
                        "Autonomous Databases"]:
        sheet = workbook.create_sheet(title=resource_type)   
        if resource_type == "Compute Instances":
            sheet.append(["Compartment", "Region", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags", "Attached_to", "BootVolume_state", "Availability_domain", "Time_created", "Total_cost_in_period"])
        if resource_type == "Block Volumes":
            sheet.append(["Compartment", "Region", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags", "Attached_to", "Size_in_gbs", "Time_created", "Total_cost_in_period"])
        if resource_type == "Block Volumes Bkp":
            sheet.append(["Compartment", "Region", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags", "Attached_to", "Size_in_gbs", "Time_created", "Total_cost_in_period"])
        if resource_type == "Boot Volumes":
            sheet.append(["Compartment", "Region", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags", "Attached_to", "Availability_domain" , "Size_in_gbs", "Time_created", "Total_cost_in_period"])
        if resource_type == "Boot Volumes Bkp":
            sheet.append(["Compartment", "Region", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags", "Size_in_gbs", "Time_created", "Total_cost_in_period"])
        if resource_type == "File Systems":
            sheet.append(["Compartment", "Region", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags" , "Metered_bytes", "Time_created", "Total_cost_in_period"])
        if resource_type == "Autonomous Databases":
            sheet.append(["Compartment", "Region", "Name", "ID", "STATE", "Defined_tags", "Freeform_tags" , "Ocpus", "Size_in_gbs", "Time_created", "Total_cost_in_period"])    
        if resource_type == "Daily Costs":
            sheet.append(["Compartment", "ID", "Currency","Cost", "Starttime"])              
        
        for compartment, resource_data in resources.items():
            for item in resource_data.get(resource_type, []):
                # Get total cost for this resource from cost_resource_ids
                resource_id = item.get("id")
                total_cost_str = item.get("total_cost_in_period", "")  # For deleted resources
                if not total_cost_str and resource_id and resource_id in cost_resource_ids:
                    cost_info = cost_resource_ids[resource_id]
                    total_cost_str = f"{cost_info['total_cost']:.4f} {cost_info['currency']}"
                           
                if resource_type == "Compute Instances":
                    sheet.append([
                             item.get("compartment_name"),
                             item.get("region"), 
                             item.get("name"), 
                             item.get("id"),
                             item.get("state"),
                             str(item.get(f"defined_tags")),
                             str(item.get(f"freeform_tags")),
                             str(item.get(f"attached_to")),
                             str(item.get(f"volume_state")),
                             str(item.get(f"availability_domain")),
                             str(item.get(f"time_created")),
                             total_cost_str
                             ])
                if resource_type == "Block Volumes":
                    sheet.append([
                                item.get("compartment_name"),
                                item.get("region"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                str(item.get(f"attached_to_instance")),
                                item.get(f"size_in_gbs"),
                                str(item.get(f"time_created")),
                                total_cost_str
                                ])
                if resource_type == "Block Volumes Bkp":
                    sheet.append([
                                item.get("compartment_name"),
                                item.get("region"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                str(item.get(f"attached_to")),
                                item.get(f"size_in_gbs"),
                                str(item.get(f"time_created")),
                                total_cost_str
                                ])
                if resource_type == "Boot Volumes":
                    sheet.append([
                                item.get("compartment_name"),
                                item.get("region"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                str(item.get(f"attached_to_instance")),
                                str(item.get(f"availability_domain")),
                                item.get(f"size_in_gbs"),
                                str(item.get(f"time_created")),
                                total_cost_str
                                ])
                if resource_type == "Boot Volumes Bkp":
                    sheet.append([
                                item.get("compartment_name"),
                                item.get("region"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                item.get(f"size_in_gbs"),
                                str(item.get(f"time_created")),
                                total_cost_str
                                ])
                if resource_type == "File Systems":
                    sheet.append([
                                item.get("compartment_name"),
                                item.get("region"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                item.get(f"metered_bytes"),
                                str(item.get(f"time_created")),
                                total_cost_str
                                ])
                if resource_type == "Autonomous Databases":
                    sheet.append([
                                item.get("compartment_name"),
                                item.get("region"), 
                                item.get("name"), 
                                item.get("id"),
                                item.get("state"),
                                str(item.get(f"defined_tags")),
                                str(item.get(f"freeform_tags")),
                                item.get(f"compute_count"),
                                item.get(f"size_in_gbs"),
                                str(item.get(f"time_created")),
                                total_cost_str
                                ])
                if resource_type == "Daily Costs":
                    sheet.append([
                                item.get("compartment_name"), 
                                item.get("id"),
                                item.get("currency"),
                                item.get("cost"),
                                item.get("starttime")
                                ])                        

    # Add visualization sheet
    visualization_sheet = workbook.create_sheet(title="Visualizations")
    visualization_sheet.append(["Resource Type", "Count"])

    # Prepare summary data for visualization
    summary_data = {}
    for compartment, resource_types in resources.items():
        for resource_type, resource_list in resource_types.items():
          if resource_type != "Daily Costs":  
            summary_data[resource_type] = summary_data.get(resource_type, 0) + len(resource_list)

    for resource_type, count in summary_data.items():
      if resource_type != "Daily Costs": 
        visualization_sheet.append([resource_type, count])

    # Create Pie Chart
    pie_chart = PieChart()
    data = Reference(visualization_sheet, min_col=2, min_row=2, max_row=len(summary_data) + 1)
    labels = Reference(visualization_sheet, min_col=1, min_row=2, max_row=len(summary_data) + 1)
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.title = "Resource Distribution"
    visualization_sheet.add_chart(pie_chart, "D2")

    # Create Bar Chart
    bar_chart = BarChart()
    bar_chart.add_data(data, titles_from_data=False)
    bar_chart.set_categories(labels)
    bar_chart.title = "Resource Counts"
    bar_chart.x_axis.title = "Resource Type"
    bar_chart.y_axis.title = "Count"
    visualization_sheet.add_chart(bar_chart, "D20")

    # Generate file name with dynamic titles
    file_name = f"oci_resources_all_regions_{namespace}_{current_date}.xlsx"
    
    # Save the Excel workbook
    workbook.save(file_name)
    print(f"Detailed findings and visualizations saved to: {file_name}")
  


except oci.exceptions.ServiceError as e:
    print(f"Service Error: {e}")
except Exception as e:
    print(f"Unexpected Error: {e}")