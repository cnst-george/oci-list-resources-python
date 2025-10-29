import oci
import json
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import PieChart, BarChart, Reference

# Load OCI configuration
configAPI = oci.config.from_file("~/.oci/config")

# Initialize OCI clients
identity_client = oci.identity.IdentityClient(configAPI)
compute_client = oci.core.ComputeClient(configAPI)
block_storage_client = oci.core.BlockstorageClient(configAPI)
file_storage_client = oci.file_storage.FileStorageClient(configAPI)
object_storage_client = oci.object_storage.ObjectStorageClient(configAPI)

# Get tenancy ID
tenancy_ocid = configAPI["tenancy"]
tenancy_name = identity_client.get_tenancy(tenancy_id=tenancy_ocid).data.name
print(f"Using Tenancy Name: {tenancy_name}")

# Get Object Storage namespace
namespace = object_storage_client.get_namespace().data

# Fetch availability domains
availability_domains = identity_client.list_availability_domains(tenancy_ocid).data

# Initialize result storage
resources = {}
findings = {}

try:
    # Fetch all compartments
    compartments = oci.pagination.list_call_get_all_results(
        identity_client.list_compartments,
        tenancy_ocid,
        compartment_id_in_subtree=True,
        access_level="ANY"
    ).data
    compartments.append(oci.identity.models.Compartment(id=tenancy_ocid, name="Tenancy Root"))

    # Discover resources in each compartment
    for compartment in compartments:
        if compartment.lifecycle_state == "ACTIVE":
            print(f"Discovering storages in compartment: {compartment.name}")
            resources[compartment.name] = {}
            findings[compartment.name] = []

            # Discover Block Volumes
            volume_response = oci.pagination.list_call_get_all_results(
                block_storage_client.list_volumes,
                compartment_id=compartment.id
            ).data
            volume_findings = []
            for volume in volume_response:
                resources[compartment.name].setdefault("Block Volumes", []).append({
                    "name": volume.display_name,
                    "id": volume.id,
                    "defined_tags" : volume.defined_tags,
                    "freeform_tags" : volume.freeform_tags
                })
                # Check if the volume is attached to any instance 
                attachments = oci.pagination.list_call_get_all_results(
                    compute_client.list_volume_attachments,
                    compartment_id=compartment.id,
                    volume_id=volume.id
                ).data
                if not attachments:  # No attachments found
                    volume_findings.append(f"Volume '{volume.display_name}' is NOT attached to any instance.")
                # Best practice: Ensure backup policy is set
                if not volume.is_auto_tune_enabled:
                    volume_findings.append(f"Volume '{volume.display_name}' does not have auto-tune enabled.")
            findings[compartment.name].extend(volume_findings)

            # Discover File Systems 
            for ad in availability_domains:  
                fss_response = oci.pagination.list_call_get_all_results(
                    file_storage_client.list_file_systems,
                    compartment_id=compartment.id,
                    availability_domain=ad.name
                ).data
                fss_findings = []
                for fss in fss_response:
                    resources[compartment.name].setdefault("File Systems", []).append({
                        "name": fss.display_name,
                        "id": fss.id,
                        "defined_tags" : fss.defined_tags,
                        "freeform_tags" : fss.freeform_tags                        
                    })
                findings[compartment.name].extend(fss_findings)

    # Get current date for the file name
    current_date = datetime.now().strftime("%Y-%m-%d")
    
    # Generate file name with dynamic titles
    file_name = f"oci_storage_{tenancy_name}_{current_date}.json"
    
    # Export data to JSON
    with open(file_name, "w") as file:
        json.dump({"resources": resources, "findings": findings}, file, indent=4)

    print(f"JSON file saved: {file_name}")
    
    # Export data to Excel
    workbook = Workbook()
    summary_sheet = workbook.active

    # Remove default sheet
    workbook.remove(workbook["Sheet"])

    # Add data sheets for each resource type
    for resource_type in ["Block Volumes"
                        ,"File Systems"
                         ]:
        sheet = workbook.create_sheet(title=resource_type)
        sheet.append(["Compartment", "Name", "ID", "Defined_tags", "Freeform_tags" ])
        for compartment, resource_data in resources.items():
            for item in resource_data.get(resource_type, []):
                sheet.append([compartment, item.get("name"), item.get("id", "N/A"),str(item.get("defined_tags")),str(item.get(f"freeform_tags"))])

      # Generate file name with dynamic titles
    file_name = f"oci_storage_{tenancy_name}_{current_date}.xlsx"
    
    # Save the Excel workbook
    workbook.save(file_name)
    print(f"Excel file saved: {file_name}")

except oci.exceptions.ServiceError as e:
    print(f"Service Error: {e}")
except Exception as e:
    print(f"Unexpected Error: {e}")